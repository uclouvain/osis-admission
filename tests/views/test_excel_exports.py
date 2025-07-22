# ##############################################################################
#
#  OSIS stands for Open Student Information System. It's an application
#  designed to manage the core business of higher education institutions,
#  such as universities, faculties, institutes and professional schools.
#  The core business involves the administration of students, teachers,
#  courses, programs and so on.
#
#  Copyright (C) 2015-2025 Université catholique de Louvain (http://www.uclouvain.be)
#
#  This program is free software: you can redistribute it and/or modify
#  it under the terms of the GNU General Public License as published by
#  the Free Software Foundation, either version 3 of the License, or
#  (at your option) any later version.
#
#  This program is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU General Public License for more details.
#
#  A copy of this license - GNU General Public License - is available
#  at the root of the source code of this program.  If not,
#  see http://www.gnu.org/licenses/.
#
# ##############################################################################

import ast
import datetime
import json
from typing import List
from unittest import mock

import freezegun
import mock
from django.conf import settings
from django.contrib.auth.models import User
from django.test import RequestFactory, TestCase, override_settings
from django.urls import reverse
from django.utils.translation import gettext as _
from django.utils.translation import pgettext, pgettext_lazy
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from osis_async.models import AsyncTask
from osis_async.models.enums import TaskState
from osis_export.models import Export
from osis_export.models.enums.types import ExportTypes

from admission.ddd.admission.doctorat.preparation.commands import (
    ListerDemandesQuery as ListerDemandesDoctoralesQuery,
)
from admission.ddd.admission.doctorat.preparation.domain.model.enums import (
    ChoixCommissionProximiteCDSS,
    ChoixStatutPropositionDoctorale,
    ChoixTypeAdmission,
    ChoixTypeFinancement,
)
from admission.ddd.admission.doctorat.preparation.dtos.liste import (
    DemandeRechercheDTO as DemandeDoctoraleRechercheDTO,
)
from admission.ddd.admission.doctorat.preparation.read_view.domain.enums.tableau_bord import (
    IndicateurTableauBordEnum,
)
from admission.ddd.admission.doctorat.validation.domain.model.enums import ChoixGenre
from admission.ddd.admission.formation_continue.domain.model.enums import (
    ChoixEdition,
    ChoixInscriptionATitre,
    ChoixMoyensDecouverteFormation,
    ChoixStatutPropositionContinue,
    ChoixTypeAdresseFacturation,
)
from admission.ddd.admission.formation_continue.domain.model.enums import (
    OngletsChecklist as OngletsChecklistContinue,
)
from admission.ddd.admission.formation_generale.domain.model.enums import (
    ChoixStatutPropositionGenerale,
)
from admission.ddd.admission.formation_generale.domain.model.enums import (
    OngletsChecklist as OngletsChecklistGenerale,
)
from admission.ddd.admission.shared_kernel.dtos.liste import (
    DemandeRechercheDTO,
    VisualiseurAdmissionDTO,
)
from admission.ddd.admission.shared_kernel.enums.checklist import ModeFiltrageChecklist
from admission.ddd.admission.shared_kernel.enums.liste import (
    TardiveModificationReorientationFiltre,
)
from admission.ddd.admission.shared_kernel.enums.type_demande import TypeDemande
from admission.ddd.admission.shared_kernel.tests.factory.profil import (
    AnneeExperienceAcademiqueDTOFactory,
    ExperienceAcademiqueDTOFactory,
    ExperienceNonAcademiqueDTOFactory,
)
from admission.tests.factories import DoctorateAdmissionFactory
from admission.tests.factories.admission_viewer import AdmissionViewerFactory
from admission.tests.factories.comment import CommentEntryFactory
from admission.tests.factories.continuing_education import (
    ContinuingEducationAdmissionFactory,
)
from admission.tests.factories.curriculum import (
    AdmissionEducationalValuatedExperiencesFactory,
    AdmissionProfessionalValuatedExperiencesFactory,
    EducationalExperienceFactory,
    EducationalExperienceYearFactory,
    ProfessionalExperienceFactory,
)
from admission.tests.factories.form_item import (
    AdmissionFormItemInstantiationFactory,
    CheckboxSelectionAdmissionFormItemFactory,
    DocumentAdmissionFormItemFactory,
    MessageAdmissionFormItemFactory,
    RadioButtonSelectionAdmissionFormItemFactory,
    TextAdmissionFormItemFactory,
)
from admission.tests.factories.general_education import GeneralEducationAdmissionFactory
from admission.tests.factories.person import CompletePersonFactory
from admission.tests.factories.roles import (
    DoctorateCommitteeMemberRoleFactory,
    SicManagementRoleFactory,
)
from admission.tests.factories.supervision import (
    CaMemberFactory,
    ExternalPromoterFactory,
    PromoterFactory,
)
from admission.views.excel_exports import (
    SPECIFIC_QUESTION_SEPARATOR,
    SPECIFIC_QUESTION_SEPARATOR_REPLACEMENT,
    AdmissionListExcelExportView,
    ContinuingAdmissionListExcelExportView,
    DoctorateAdmissionListExcelExportView,
)
from base.models.enums.civil_state import CivilState
from base.models.enums.education_group_types import TrainingType
from base.models.enums.entity_type import EntityType
from base.models.enums.got_diploma import GotDiploma
from base.tests import QueriesAssertionsMixin
from base.tests.factories.academic_year import AcademicYearFactory
from base.tests.factories.campus import CampusFactory
from base.tests.factories.entity import EntityFactory
from base.tests.factories.entity_version import EntityVersionFactory
from base.tests.factories.organization import OrganizationFactory
from base.tests.factories.person import PersonFactory
from base.tests.factories.student import StudentFactory
from epc.models.enums.decision_resultat_cycle import DecisionResultatCycle
from epc.tests.factories.inscription_programme_annuel import (
    InscriptionProgrammeAnnuelFactory,
)
from infrastructure.messages_bus import message_bus_instance
from osis_profile.models import EducationalExperience, ProfessionalExperience
from osis_profile.models.enums.curriculum import ActivitySector, ActivityType
from program_management.models.education_group_version import EducationGroupVersion
from reference.models.enums.cycle import Cycle
from reference.tests.factories.country import CountryFactory
from reference.tests.factories.diploma_title import DiplomaTitleFactory
from reference.tests.factories.scholarship import (
    DoctorateScholarshipFactory,
    DoubleDegreeScholarshipFactory,
    ErasmusMundusScholarshipFactory,
    InternationalScholarshipFactory,
)


class UnfrozenDTO:
    # Trick to make this "unfrozen" just for tests
    def __setattr__(self, key, value):
        object.__setattr__(self, key, value)

    def __delattr__(self, item):
        object.__delattr__(self, item)


class _DemandeRechercheDTO(UnfrozenDTO, DemandeRechercheDTO):
    pass


@freezegun.freeze_time('2023-01-01')
@override_settings(WAFFLE_CREATE_MISSING_SWITCHES=False)
class AdmissionListExcelExportViewTestCase(QueriesAssertionsMixin, TestCase):
    def assertStrEqual(self, first, second, msg=None):
        self.assertIsInstance(first, str)
        self.assertIsInstance(second, str)
        super().assertEqual(first, second, msg)

    @classmethod
    def setUpTestData(cls):
        cls.factory = RequestFactory()

        # Users
        cls.user = User.objects.create_user(
            username='john_doe',
            password='top_secret',
        )

        cls.sic_management_user = SicManagementRoleFactory().person.user
        cls.other_sic_manager = SicManagementRoleFactory().person

        # Academic years
        AcademicYearFactory.produce(base_year=2023, number_past=2, number_future=2)

        # Entities
        faculty_entity = EntityFactory()
        EntityVersionFactory(
            entity=faculty_entity,
            acronym='ABCDEF',
            entity_type=EntityType.FACULTY.name,
        )

        cls.first_entity = EntityFactory()
        EntityVersionFactory(
            entity=cls.first_entity,
            acronym='GHIJK',
            entity_type=EntityType.SCHOOL.name,
            parent=faculty_entity,
        )

        # Admissions
        cls.admission = GeneralEducationAdmissionFactory(
            candidate__country_of_citizenship=CountryFactory(european_union=True, name='Belgique'),
            candidate__first_name="John",
            candidate__last_name="Doe",
            candidate__private_email="jdoe@example.be",
            status=ChoixStatutPropositionGenerale.CONFIRMEE.name,
            training__management_entity=cls.first_entity,
            training__acronym="ABCD0",
            last_update_author__user__username='user1',
            submitted_at=datetime.date(2023, 1, 2),
        )

        cls.admission.last_update_author = cls.admission.candidate
        cls.admission.save()

        admission_viewers = [
            AdmissionViewerFactory(person=cls.sic_management_user.person, admission=cls.admission),
            AdmissionViewerFactory(person=cls.other_sic_manager, admission=cls.admission),
        ]

        cls.lite_reference = '{:07,}'.format(cls.admission.reference).replace(',', '.')

        cls.student = StudentFactory(
            person=cls.admission.candidate,
            registration_id='01234567',
        )

        teaching_campus = (
            EducationGroupVersion.objects.filter(offer=cls.admission.training)
            .first()
            .root_group.main_teaching_campus.name
        )

        cls.message_form_item = AdmissionFormItemInstantiationFactory(
            form_item=MessageAdmissionFormItemFactory(
                internal_label='Q1',
            ),
            academic_year=cls.admission.determined_academic_year,
        ).form_item
        cls.message_form_item_uuid = str(cls.message_form_item.uuid)

        cls.text_form_item = AdmissionFormItemInstantiationFactory(
            form_item=TextAdmissionFormItemFactory(
                internal_label='Q2',
            ),
            academic_year=cls.admission.determined_academic_year,
        ).form_item
        cls.text_form_item_uuid = str(cls.text_form_item.uuid)

        cls.inactive_text_form_item = AdmissionFormItemInstantiationFactory(
            form_item=TextAdmissionFormItemFactory(
                internal_label='Q3',
                active=False,
            ),
            academic_year=cls.admission.determined_academic_year,
        ).form_item
        cls.inactive_text_form_item_uuid = str(cls.inactive_text_form_item.uuid)

        cls.document_form_item = AdmissionFormItemInstantiationFactory(
            form_item=DocumentAdmissionFormItemFactory(
                internal_label='Q4',
            ),
            academic_year=cls.admission.determined_academic_year,
        ).form_item
        cls.document_form_item_uuid = str(cls.document_form_item.uuid)

        cls.checkbox_form_item = AdmissionFormItemInstantiationFactory(
            form_item=CheckboxSelectionAdmissionFormItemFactory(
                internal_label='Q5',
            ),
            academic_year=cls.admission.determined_academic_year,
        ).form_item
        cls.checkbox_form_item_uuid = str(cls.checkbox_form_item.uuid)

        cls.radio_button_form_item = AdmissionFormItemInstantiationFactory(
            form_item=RadioButtonSelectionAdmissionFormItemFactory(
                internal_label='Q6',
                values=[
                    {
                        'key': '1',
                        'en': f'One{SPECIFIC_QUESTION_SEPARATOR}A',
                        'fr-be': f'Un{SPECIFIC_QUESTION_SEPARATOR}A',
                    },
                    {'key': '2', 'en': 'Two', 'fr-be': 'Deux'},
                    {'key': '3', 'en': 'Three', 'fr-be': 'Trois'},
                ],
            ),
            academic_year=cls.admission.determined_academic_year,
        ).form_item
        cls.radio_button_form_item_uuid = str(cls.radio_button_form_item.uuid)

        cls.result = _DemandeRechercheDTO(
            uuid=cls.admission.uuid,
            numero_demande=f'M-ABCDEF22-{cls.lite_reference}',
            nom_candidat=cls.admission.candidate.last_name,
            prenom_candidat=cls.admission.candidate.first_name,
            noma_candidat=cls.admission.candidate.last_registration_id,
            plusieurs_demandes=False,
            sigle_formation=cls.admission.training.acronym,
            code_formation=cls.admission.training.partial_acronym,
            intitule_formation=cls.admission.training.title,
            type_formation=cls.admission.training.education_group_type.name,
            lieu_formation=teaching_campus,
            est_inscription_tardive=None,
            est_modification_inscription_externe=None,
            est_reorientation_inscription_externe=None,
            nationalite_candidat=cls.admission.candidate.country_of_citizenship.name,
            nationalite_ue_candidat=cls.admission.candidate.country_of_citizenship.european_union,
            vip=any(
                scholarship
                for scholarship in [
                    cls.admission.erasmus_mundus_scholarship,
                    cls.admission.international_scholarship,
                    cls.admission.double_degree_scholarship,
                ]
            ),
            etat_demande=cls.admission.status,
            type_demande=cls.admission.type_demande,
            derniere_modification_le=cls.admission.modified_at,
            derniere_modification_par=cls.admission.last_update_author.user.username,
            derniere_modification_par_candidat=True,
            dernieres_vues_par=[
                VisualiseurAdmissionDTO(
                    nom=admission_viewers[1].person.last_name,
                    prenom=admission_viewers[1].person.first_name,
                    date=admission_viewers[1].viewed_at,
                ),
            ],
            date_confirmation=cls.admission.submitted_at,
            est_premiere_annee=False,
            poursuite_de_cycle=cls.admission.cycle_pursuit,
            annee_formation=cls.admission.training.academic_year.year,
            annee_calculee=(
                cls.admission.determined_academic_year.year if cls.admission.determined_academic_year else None
            ),
            adresse_email_candidat=cls.admission.candidate.private_email,
            reponses_questions_specifiques={
                cls.text_form_item_uuid: 'Answer 1',
                cls.inactive_text_form_item_uuid: 'Answer 2',
                cls.checkbox_form_item_uuid: ['1', '2'],
                cls.radio_button_form_item_uuid: '3',
            },
        )

        cls.default_params = {
            'annee_academique': 2022,
            'taille_page': 10,
            'demandeur': str(cls.sic_management_user.person.uuid),
        }

        # Targeted url
        cls.url = reverse('admission:excel-exports:all-admissions-list')
        cls.list_url = reverse('admission:all-list')

    def test_export_user_without_person(self):
        self.client.force_login(user=self.user)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 403)

    def test_export_candidate_user(self):
        self.client.force_login(user=self.admission.candidate.user)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 403)

    def test_export_sic_management_user(self):
        self.client.force_login(user=self.sic_management_user)

        response = self.client.get(self.url)
        self.assertRedirects(response, expected_url=self.list_url, fetch_redirect_response=False)

        response = self.client.get(
            self.url,
            **{"HTTP_HX-Request": 'true'},
        )

        self.assertEqual(response.status_code, 200)

    def test_export_with_sic_management_user_without_filters_doesnt_plan_the_export(self):
        self.client.force_login(user=self.sic_management_user)

        response = self.client.get(self.url)

        self.assertRedirects(response, expected_url=self.list_url, fetch_redirect_response=False)

        task = AsyncTask.objects.filter(person=self.sic_management_user.person).first()
        self.assertIsNone(task)

        export = Export.objects.filter(person=self.sic_management_user.person).first()
        self.assertIsNone(export)

    def test_export_with_sic_management_user_with_filters_plans_the_export(self):
        self.client.force_login(user=self.sic_management_user)

        response = self.client.get(self.url, data=self.default_params)

        self.assertRedirects(response, expected_url=self.list_url, fetch_redirect_response=False)

        task = AsyncTask.objects.filter(person=self.sic_management_user.person).first()
        self.assertIsNotNone(task)
        self.assertEqual(task.name, _('Admission applications export'))
        self.assertEqual(task.description, _('Excel export of admission applications'))
        self.assertEqual(task.state, TaskState.PENDING.name)

        export = Export.objects.filter(job_uuid=task.uuid).first()
        self.assertIsNotNone(export)
        self.assertEqual(export.called_from_class, 'admission.views.excel_exports.AdmissionListExcelExportView')
        self.assertEqual(export.file_name, 'export-des-demandes-dadmission')
        self.assertEqual(export.type, ExportTypes.EXCEL.name)

        filters = ast.literal_eval(export.filters)
        self.assertEqual(filters.get('annee_academique'), self.default_params.get('annee_academique'))
        self.assertEqual(filters.get('demandeur'), self.default_params.get('demandeur'))

    def test_export_with_sic_management_user_with_filters_and_asc_ordering(self):
        self.client.force_login(user=self.sic_management_user)

        # With asc ordering
        response = self.client.get(self.url, data={**self.default_params, 'o': 'numero_demande'})

        self.assertRedirects(response, expected_url=self.list_url, fetch_redirect_response=False)

        task = AsyncTask.objects.filter(person=self.sic_management_user.person).first()
        self.assertIsNotNone(task)

        export = Export.objects.filter(job_uuid=task.uuid).first()
        self.assertIsNotNone(export)

        filters = ast.literal_eval(export.filters)
        self.assertEqual(filters.get('annee_academique'), self.default_params.get('annee_academique'))
        self.assertEqual(filters.get('demandeur'), self.default_params.get('demandeur'))
        self.assertEqual(filters.get('tri_inverse'), False)
        self.assertEqual(filters.get('champ_tri'), 'numero_demande')

    def test_export_with_sic_management_user_with_filters_and_desc_ordering(self):
        self.client.force_login(user=self.sic_management_user)

        # With asc ordering
        response = self.client.get(self.url, data={**self.default_params, 'o': '-numero_demande'})

        self.assertRedirects(response, expected_url=self.list_url, fetch_redirect_response=False)

        task = AsyncTask.objects.filter(person=self.sic_management_user.person).first()
        self.assertIsNotNone(task)

        export = Export.objects.filter(job_uuid=task.uuid).first()
        self.assertIsNotNone(export)

        filters = ast.literal_eval(export.filters)
        self.assertEqual(filters.get('annee_academique'), self.default_params.get('annee_academique'))
        self.assertEqual(filters.get('demandeur'), self.default_params.get('demandeur'))
        self.assertEqual(filters.get('tri_inverse'), True)
        self.assertEqual(filters.get('champ_tri'), 'numero_demande')

    def test_export_content(self):
        view = AdmissionListExcelExportView()
        view.initialize_specific_questions()
        header = view.get_header()
        row_data = view.get_row_data(self.result)
        self.assertEqual(len(header), len(row_data))

        self.assertStrEqual(row_data[0], self.result.numero_demande)
        self.assertStrEqual(row_data[1], self.result.nom_candidat)
        self.assertStrEqual(row_data[2], self.result.prenom_candidat)
        self.assertStrEqual(row_data[3], self.result.noma_candidat)
        self.assertStrEqual(row_data[4], 'non')
        self.assertStrEqual(row_data[5], self.result.sigle_formation)
        self.assertStrEqual(row_data[6], self.result.intitule_formation)
        self.assertStrEqual(row_data[7], self.result.nationalite_candidat)
        self.assertStrEqual(row_data[8], 'oui')
        self.assertStrEqual(row_data[9], str(ChoixStatutPropositionGenerale.CONFIRMEE.value))
        self.assertStrEqual(row_data[10], _('candidate'))
        self.assertStrEqual(row_data[11], '2023/01/01, 00:00:00')
        self.assertStrEqual(row_data[12], '2023/01/02, 00:00:00')
        self.assertStrEqual(row_data[13], self.result.adresse_email_candidat)
        answers_to_specific_questions = row_data[14].split(SPECIFIC_QUESTION_SEPARATOR)
        self.assertCountEqual(
            answers_to_specific_questions,
            ['Q2 : Answer 1', 'Q5 : Un,Deux', 'Q6 : Trois'],
        )

        with mock.patch.object(self.result, 'date_confirmation', None):
            row_data = view.get_row_data(self.result)
            self.assertStrEqual(row_data[12], '')

        with mock.patch.object(self.result, 'plusieurs_demandes', True):
            row_data = view.get_row_data(self.result)
            self.assertStrEqual(row_data[4], 'oui')

    def test_export_content_with_invalid_specific_select_question_answer(self):
        view = AdmissionListExcelExportView()
        view.initialize_specific_questions()
        view.language = settings.LANGUAGE_CODE_FR

        with mock.patch.object(
            self.result,
            'reponses_questions_specifiques',
            {
                self.checkbox_form_item_uuid: ['1', '8'],
                self.radio_button_form_item_uuid: '8',
            },
        ):
            row_data = view.get_row_data(self.result)
            answers_to_specific_questions = row_data[14].split(SPECIFIC_QUESTION_SEPARATOR)
            self.assertCountEqual(
                answers_to_specific_questions,
                ['Q5 : Un', 'Q6 : '],
            )

    def test_export_content_with_specific_questions_answers_containing_the_separator(self):
        view = AdmissionListExcelExportView()
        view.initialize_specific_questions()
        view.language = settings.LANGUAGE_CODE_EN

        with mock.patch.object(
            self.result,
            'reponses_questions_specifiques',
            {
                self.text_form_item_uuid: f'A{SPECIFIC_QUESTION_SEPARATOR}B',
                self.radio_button_form_item_uuid: '1',
            },
        ):
            row_data = view.get_row_data(self.result)
            answers_to_specific_questions = row_data[14].split(SPECIFIC_QUESTION_SEPARATOR)
            self.assertCountEqual(
                answers_to_specific_questions,
                [
                    f'Q2 : A{SPECIFIC_QUESTION_SEPARATOR_REPLACEMENT}B',
                    f'Q6 : One{SPECIFIC_QUESTION_SEPARATOR_REPLACEMENT}A',
                ],
            )

    def test_export_configuration(self):
        candidate = PersonFactory()
        campus = CampusFactory()
        international_scholarship = InternationalScholarshipFactory(short_name='ID1')
        double_degree_scholarship = DoubleDegreeScholarshipFactory(short_name="DD1")
        erasmus_mundus_scholarship = ErasmusMundusScholarshipFactory(short_name="EM1")
        filters_nb = 21
        filters = {
            'annee_academique': 2022,
            'numero': 1,
            'noma': '00000001',
            'matricule_candidat': candidate.global_id,
            'etats': [ChoixStatutPropositionGenerale.CONFIRMEE.name],
            'type': TypeDemande.ADMISSION.name,
            'site_inscription': str(campus.uuid),
            'entites': 'ENT',
            'types_formation': [TrainingType.BACHELOR.name, TrainingType.PHD.name],
            'formation': 'Informatique',
            'bourse_internationale': str(international_scholarship.uuid),
            'bourse_erasmus_mundus': str(erasmus_mundus_scholarship.uuid),
            'bourse_double_diplomation': str(double_degree_scholarship.uuid),
            'demandeur': str(self.sic_management_user.person.uuid),
            'mode_filtres_etats_checklist': ModeFiltrageChecklist.INCLUSION.name,
            'filtres_etats_checklist': {
                OngletsChecklistGenerale.donnees_personnelles.name: ['A_TRAITER'],
                OngletsChecklistGenerale.frais_dossier.name: ['PAYES'],
            },
            'quarantaine': 'True',
            'tardif_modif_reorientation': TardiveModificationReorientationFiltre.INSCRIPTION_TARDIVE.name,
            'delai_depasse_complements': 'True',
        }

        view = AdmissionListExcelExportView()
        workbook = Workbook()
        worksheet: Worksheet = workbook.create_sheet()

        view.customize_parameters_worksheet(
            worksheet=worksheet,
            person=self.sic_management_user.person,
            filters=str(filters),
        )

        names, values = list(worksheet.iter_cols(values_only=True))
        self.assertEqual(len(names), filters_nb)
        self.assertEqual(len(values), filters_nb)

        # Check the names of the parameters
        self.assertStrEqual(names[0], _('Creation date'))
        self.assertStrEqual(names[1], pgettext('masculine', 'Created by'))
        self.assertStrEqual(names[2], _('Description'))
        self.assertStrEqual(names[3], _('Year'))
        self.assertStrEqual(names[4], _('Application numero'))
        self.assertStrEqual(names[5], _('Noma'))
        self.assertStrEqual(names[6], _('Last name / First name / Email'))
        self.assertStrEqual(names[7], _('Application status'))
        self.assertStrEqual(names[8], _('Application type'))
        self.assertStrEqual(names[9], _('Enrolment campus'))
        self.assertStrEqual(names[10], pgettext('admission', 'Entities'))
        self.assertStrEqual(names[11], _('Course type'))
        self.assertStrEqual(names[12], pgettext('admission', 'Course'))
        self.assertStrEqual(names[13], _('International scholarship'))
        self.assertStrEqual(names[14], _('Erasmus Mundus'))
        self.assertStrEqual(names[15], _('Dual degree scholarship'))
        self.assertStrEqual(names[16], _('Include or exclude the checklist filters'))
        self.assertStrEqual(names[17], _('Checklist filters'))
        self.assertStrEqual(names[18], _('Quarantine'))
        self.assertStrEqual(names[19], _('Late/Modif./Reor.'))
        self.assertStrEqual(names[20], _('Deadline for complements'))

        # Check the values of the parameters
        self.assertStrEqual(values[0], '1 Janvier 2023')
        self.assertStrEqual(values[1], self.sic_management_user.person.full_name)
        self.assertStrEqual(values[2], _('Export') + ' - Admissions')
        self.assertStrEqual(values[3], '2022')
        self.assertStrEqual(values[4], '1')
        self.assertStrEqual(values[5], '00000001')
        self.assertStrEqual(values[6], candidate.full_name)
        self.assertStrEqual(values[7], f"['{ChoixStatutPropositionGenerale.CONFIRMEE.value}']")
        self.assertStrEqual(values[8], str(TypeDemande.ADMISSION.value))
        self.assertStrEqual(values[9], campus.name)
        self.assertStrEqual(values[10], 'ENT')
        self.assertStrEqual(values[11], f"['{TrainingType.BACHELOR.value}', '{TrainingType.PHD.value}']")
        self.assertStrEqual(values[12], 'Informatique')
        self.assertStrEqual(values[13], international_scholarship.short_name)
        self.assertStrEqual(values[14], erasmus_mundus_scholarship.short_name)
        self.assertStrEqual(values[15], double_degree_scholarship.short_name)
        self.assertStrEqual(values[16], str(ModeFiltrageChecklist.INCLUSION.value))
        self.assertStrEqual(
            values[17],
            str(
                {
                    OngletsChecklistGenerale.donnees_personnelles.value: [_('To be processed')],
                    OngletsChecklistGenerale.frais_dossier.value: [_('Payed')],
                }
            ),
        )
        self.assertStrEqual(values[18], 'Oui')
        self.assertStrEqual(values[19], 'Inscription tardive')
        self.assertStrEqual(values[20], _('Deadline exceeded'))

        filters['quarantaine'] = False

        worksheet: Worksheet = workbook.create_sheet()

        view.customize_parameters_worksheet(
            worksheet=worksheet,
            person=self.sic_management_user.person,
            filters=str(filters),
        )

        names, values = list(worksheet.iter_cols(values_only=True))
        self.assertEqual(len(names), filters_nb)
        self.assertEqual(len(values), filters_nb)

        self.assertStrEqual(values[18], 'Non')

        filters['quarantaine'] = None

        worksheet: Worksheet = workbook.create_sheet()

        view.customize_parameters_worksheet(
            worksheet=worksheet,
            person=self.sic_management_user.person,
            filters=str(filters),
        )

        names, values = list(worksheet.iter_cols(values_only=True))
        self.assertEqual(len(names), filters_nb)
        self.assertEqual(len(values), filters_nb)

        self.assertStrEqual(values[18], 'Tous')

        filters['delai_depasse_complements'] = ''

        worksheet: Worksheet = workbook.create_sheet()

        view.customize_parameters_worksheet(
            worksheet=worksheet,
            person=self.sic_management_user.person,
            filters=str(filters),
        )

        names, values = list(worksheet.iter_cols(values_only=True))
        self.assertEqual(len(names), filters_nb)
        self.assertEqual(len(values), filters_nb)

        self.assertStrEqual(values[20], '')


@freezegun.freeze_time('2023-01-03')
@override_settings(WAFFLE_CREATE_MISSING_SWITCHES=False)
class ContinuingAdmissionListExcelExportViewTestCase(QueriesAssertionsMixin, TestCase):
    def assertStrEqual(self, first, second, msg=None):
        self.assertIsInstance(first, str)
        self.assertIsInstance(second, str)
        super().assertEqual(first, second, msg)

    @classmethod
    def setUpTestData(cls):
        cls.factory = RequestFactory()

        # Users
        cls.user = User.objects.create_user(
            username='john_doe',
            password='top_secret',
        )

        cls.sic_management_user = SicManagementRoleFactory().person.user
        cls.other_sic_manager = SicManagementRoleFactory().person

        # Academic years
        academic_years = AcademicYearFactory.produce(base_year=2023, number_past=2, number_future=2)

        # Entities
        faculty_entity = EntityFactory()
        EntityVersionFactory(
            entity=faculty_entity,
            acronym='ABCDEF',
            entity_type=EntityType.FACULTY.name,
        )

        cls.first_entity = EntityFactory()
        EntityVersionFactory(
            entity=cls.first_entity,
            acronym='GHIJK',
            entity_type=EntityType.SCHOOL.name,
            parent=faculty_entity,
        )

        # Admissions
        candidate = CompletePersonFactory(
            first_name="John",
            last_name="Doe",
            email='john.doe@example.be',
            private_email='john.doe@example.private.be',
            gender=ChoixGenre.H.name,
            civil_state=CivilState.LEGAL_COHABITANT.name,
            country_of_citizenship=CountryFactory(name='B1'),
            birth_date=datetime.date(2020, 1, 1),
            birth_place='Place 1',
            birth_country=CountryFactory(name='B2'),
            national_number='N1',
            id_card_number='N2',
            passport_number='N3',
            last_registration_year=AcademicYearFactory(year=2015),
            last_registration_id='NOMA1',
            emergency_contact_phone='010203',
            phone_mobile='01020304',
            graduated_from_high_school=GotDiploma.YES.name,
            graduated_from_high_school_year=AcademicYearFactory(year=2014),
        )
        experiences = EducationalExperience.objects.filter(person=candidate)
        experience = experiences[0]
        experience.obtained_diploma = True
        experience.education_name = ('',)
        experience.institute_name = ('',)
        experience.institute = OrganizationFactory(
            name='I1',
        )
        experience.program = DiplomaTitleFactory(
            title='Computer science 1',
            cycle=Cycle.FIRST_CYCLE.name,
        )
        experience.save()

        past_non_academic_activity = ProfessionalExperience.objects.filter(
            person=candidate,
        ).first()
        past_non_academic_activity.type = ActivityType.INTERNSHIP.name
        past_non_academic_activity.save()

        current_non_academic_activity = ProfessionalExperienceFactory(
            person=candidate,
            start_date=datetime.date(2022, 1, 1),
            end_date=datetime.date(2023, 4, 1),
            type=ActivityType.WORK.name,
            role='R1',
            sector=ActivitySector.ASSOCIATIVE.name,
            institute_name='IA1',
        )
        cls.address_country_name = candidate.personaddress_set.first().country.name
        cls.admission = ContinuingEducationAdmissionFactory(
            candidate=candidate,
            status=ChoixStatutPropositionContinue.CONFIRMEE.name,
            training__management_entity=cls.first_entity,
            training__acronym="ZEBU0",
            training__education_group_type__name=TrainingType.UNIVERSITY_FIRST_CYCLE_CERTIFICATE.name,
            submitted_at=datetime.datetime(2023, 1, 1),
            training__academic_year=academic_years[1],
            determined_academic_year=academic_years[2],
            edition=ChoixEdition.DEUX.name,
            modified_at=datetime.datetime(2023, 1, 3),
            last_update_author=candidate,
            head_office_name='HO1',
            unique_business_number='U1',
            vat_number='VAT1',
            billing_address_type=ChoixTypeAdresseFacturation.AUTRE.name,
            billing_address_street='School street',
            billing_address_street_number='12',
            billing_address_postal_code='1000',
            billing_address_city='Bruxelles',
            billing_address_country=CountryFactory(name='BE3'),
            in_payement_order=False,
            reduced_rights=True,
            pay_by_training_cheque=None,
            cep=False,
            payement_spread=True,
            training_spread=None,
            experience_knowledge_valorisation=False,
            assessment_test_presented=True,
            assessment_test_succeeded=None,
            certificate_provided=False,
            tff_label='TFF',
            ways_to_find_out_about_the_course=[
                ChoixMoyensDecouverteFormation.ANCIENS_ETUDIANTS.name,
                ChoixMoyensDecouverteFormation.LINKEDIN.name,
            ],
            other_way_to_find_out_about_the_course='Other way',
        )

        cls.text_form_item = AdmissionFormItemInstantiationFactory(
            form_item=TextAdmissionFormItemFactory(
                internal_label='Q2',
                active=True,
            ),
            academic_year=cls.admission.determined_academic_year,
        ).form_item
        cls.text_form_item_uuid = str(cls.text_form_item.uuid)

        cls.admission.specific_question_answers = {
            cls.text_form_item_uuid: 'T1',
        }
        cls.admission.save()

        CommentEntryFactory(
            object_uuid=cls.admission.uuid,
            content='TEST',
            tags=[OngletsChecklistContinue.fiche_etudiant.name],
        )

        AdmissionEducationalValuatedExperiencesFactory(
            baseadmission=cls.admission,
            educationalexperience=experience,
        )

        AdmissionProfessionalValuatedExperiencesFactory(
            baseadmission=cls.admission,
            professionalexperience=past_non_academic_activity,
        )

        AdmissionProfessionalValuatedExperiencesFactory(
            baseadmission=cls.admission,
            professionalexperience=current_non_academic_activity,
        )

        cls.student = StudentFactory(
            person=candidate,
            registration_id='01234567',
        )

        cls.default_params = {
            'annee_academique': 2022,
            'taille_page': 10,
            'demandeur': str(cls.sic_management_user.person.uuid),
        }

        # Targeted url
        cls.url = reverse('admission:excel-exports:continuing-admissions-list')
        cls.list_url = reverse('admission:continuing-education:list')

    def test_export_user_without_person(self):
        self.client.force_login(user=self.user)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 403)

    def test_export_candidate_user(self):
        self.client.force_login(user=self.admission.candidate.user)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 403)

    def test_export_sic_management_user(self):
        self.client.force_login(user=self.sic_management_user)

        response = self.client.get(self.url)
        self.assertRedirects(response, expected_url=self.list_url, fetch_redirect_response=False)

        response = self.client.get(
            self.url,
            **{"HTTP_HX-Request": 'true'},
        )

        self.assertEqual(response.status_code, 200)

    def test_export_with_sic_management_user_without_filters_doesnt_plan_the_export(self):
        self.client.force_login(user=self.sic_management_user)

        response = self.client.get(self.url)

        self.assertRedirects(response, expected_url=self.list_url, fetch_redirect_response=False)

        task = AsyncTask.objects.filter(person=self.sic_management_user.person).first()
        self.assertIsNone(task)

        export = Export.objects.filter(person=self.sic_management_user.person).first()
        self.assertIsNone(export)

    def test_export_with_sic_management_user_with_filters_plans_the_export(self):
        self.client.force_login(user=self.sic_management_user)

        response = self.client.get(self.url, data=self.default_params)

        self.assertRedirects(response, expected_url=self.list_url, fetch_redirect_response=False)

        task = AsyncTask.objects.filter(person=self.sic_management_user.person).first()
        self.assertIsNotNone(task)
        self.assertEqual(task.name, _('Admission applications export'))
        self.assertEqual(task.description, _('Excel export of admission applications'))
        self.assertEqual(task.state, TaskState.PENDING.name)

        export = Export.objects.filter(job_uuid=task.uuid).first()
        self.assertIsNotNone(export)
        self.assertEqual(
            export.called_from_class,
            'admission.views.excel_exports.ContinuingAdmissionListExcelExportView',
        )
        self.assertEqual(export.file_name, 'export-des-demandes-dadmission')
        self.assertEqual(export.type, ExportTypes.EXCEL.name)

        filters = ast.literal_eval(export.filters)
        self.assertEqual(filters.get('annee_academique'), self.default_params.get('annee_academique'))
        self.assertEqual(filters.get('demandeur'), self.default_params.get('demandeur'))

    def test_export_with_sic_management_user_with_filters_and_asc_ordering(self):
        self.client.force_login(user=self.sic_management_user)

        # With asc ordering
        response = self.client.get(self.url, data={**self.default_params, 'o': 'numero_demande'})

        self.assertRedirects(response, expected_url=self.list_url, fetch_redirect_response=False)

        task = AsyncTask.objects.filter(person=self.sic_management_user.person).first()
        self.assertIsNotNone(task)

        export = Export.objects.filter(job_uuid=task.uuid).first()
        self.assertIsNotNone(export)

        filters = ast.literal_eval(export.filters)
        self.assertEqual(filters.get('annee_academique'), self.default_params.get('annee_academique'))
        self.assertEqual(filters.get('demandeur'), self.default_params.get('demandeur'))
        self.assertEqual(filters.get('tri_inverse'), False)
        self.assertEqual(filters.get('champ_tri'), 'numero_demande')

    def test_export_with_sic_management_user_with_filters_and_desc_ordering(self):
        self.client.force_login(user=self.sic_management_user)

        # With asc ordering
        response = self.client.get(self.url, data={**self.default_params, 'o': '-numero_demande'})

        self.assertRedirects(response, expected_url=self.list_url, fetch_redirect_response=False)

        task = AsyncTask.objects.filter(person=self.sic_management_user.person).first()
        self.assertIsNotNone(task)

        export = Export.objects.filter(job_uuid=task.uuid).first()
        self.assertIsNotNone(export)

        filters = ast.literal_eval(export.filters)
        self.assertEqual(filters.get('annee_academique'), self.default_params.get('annee_academique'))
        self.assertEqual(filters.get('demandeur'), self.default_params.get('demandeur'))
        self.assertEqual(filters.get('tri_inverse'), True)
        self.assertEqual(filters.get('champ_tri'), 'numero_demande')

    def test_export_content(self):
        view = ContinuingAdmissionListExcelExportView()
        view.initialize_specific_questions()
        header = view.get_header()

        row_data = view.get_row_data(self.admission.uuid)

        self.assertEqual(len(header), len(row_data))

        self.assertStrEqual(row_data[0], 'Doe')
        self.assertStrEqual(row_data[1], 'John')
        self.assertStrEqual(row_data[2], 'john.doe@example.private.be')
        self.assertStrEqual(row_data[3], str(ChoixStatutPropositionContinue.CONFIRMEE.value))
        self.assertStrEqual(row_data[4], str(ChoixGenre.H.value))
        self.assertStrEqual(row_data[5], str(CivilState.LEGAL_COHABITANT.value))
        self.assertStrEqual(row_data[6], 'B1')
        self.assertStrEqual(row_data[7], '2020-01-01'),
        self.assertStrEqual(row_data[8], 'Place 1')
        self.assertStrEqual(row_data[9], 'B2')
        self.assertStrEqual(row_data[10], 'N1')
        self.assertStrEqual(row_data[11], 'N2')
        self.assertStrEqual(row_data[12], 'N3')
        self.assertStrEqual(row_data[13], 'oui')
        self.assertStrEqual(row_data[14], '2015-2016')
        self.assertStrEqual(row_data[15], 'NOMA1')
        self.assertStrEqual(row_data[16], f'University street 1 - 1348 Louvain-la-Neuve - {self.address_country_name}')
        self.assertStrEqual(row_data[17], f'University street 2 - 1348 Louvain-la-Neuve - {self.address_country_name}')
        self.assertStrEqual(row_data[18], '010203')
        self.assertStrEqual(row_data[19], '01020304')
        self.assertStrEqual(row_data[20], str(GotDiploma.YES.value))
        self.assertEqual(row_data[21], 2015)
        self.assertStrEqual(row_data[22], str(Cycle.FIRST_CYCLE.value))
        self.assertStrEqual(row_data[23], 'I1')
        self.assertEqual(row_data[24], 2023)
        self.assertStrEqual(row_data[25], '')
        self.assertStrEqual(row_data[26], 'R1')
        self.assertStrEqual(row_data[27], 'IA1')
        self.assertStrEqual(row_data[28], str(ActivitySector.ASSOCIATIVE.value))
        self.assertStrEqual(row_data[29], f'01/2021-03/2021 : {ActivityType.INTERNSHIP.value}')
        self.assertStrEqual(row_data[30], 'My motivations')
        self.assertStrEqual(row_data[31], self.admission.training.acronym)
        self.assertStrEqual(row_data[32], 'ABCDEF')
        self.assertStrEqual(row_data[33], '')
        self.assertStrEqual(
            row_data[34],
            '{}\n{}\n{}'.format(
                ChoixMoyensDecouverteFormation.ANCIENS_ETUDIANTS.value,
                ChoixMoyensDecouverteFormation.LINKEDIN.value,
                'Other way',
            ),
        )
        self.assertStrEqual(row_data[35], str(ChoixInscriptionATitre.PRIVE.value))
        self.assertStrEqual(row_data[36], 'HO1')
        self.assertStrEqual(row_data[37], 'U1')
        self.assertStrEqual(row_data[38], 'VAT1')
        self.assertStrEqual(row_data[39], 'School street 12 - 1000 Bruxelles - BE3')
        self.assertStrEqual(row_data[40], 'Q2 : T1')
        self.assertStrEqual(row_data[41], 'non')
        self.assertStrEqual(row_data[42], 'oui')
        self.assertStrEqual(row_data[43], 'non')
        self.assertStrEqual(row_data[44], 'oui')
        self.assertStrEqual(row_data[45], '')
        self.assertStrEqual(row_data[46], 'non')
        self.assertStrEqual(row_data[47], 'oui')
        self.assertStrEqual(row_data[48], '')
        self.assertStrEqual(row_data[49], 'non')
        self.assertStrEqual(row_data[50], 'oui')
        self.assertStrEqual(row_data[51], '')
        self.assertStrEqual(row_data[52], 'TFF')
        self.assertStrEqual(row_data[53], 'non')
        self.assertStrEqual(row_data[54], 'TEST')

        self.admission.billing_address_type = ChoixTypeAdresseFacturation.CONTACT.name
        self.admission.save()

        self.admission.candidate.last_registration_year = None
        self.admission.candidate.graduated_from_high_school = GotDiploma.NO.name
        self.admission.candidate.graduated_from_high_school_year = None

        self.admission.candidate.save()

        new_experience = EducationalExperienceFactory(
            person=self.admission.candidate,
            education_name='',
            program=DiplomaTitleFactory(
                title='Computer science 2',
                cycle=Cycle.SECOND_CYCLE.name,
            ),
            obtained_diploma=True,
        )
        new_experience_year = EducationalExperienceYearFactory(
            academic_year__year=2010,
            educational_experience=new_experience,
        )
        AdmissionEducationalValuatedExperiencesFactory(
            baseadmission=self.admission,
            educationalexperience=new_experience,
        )

        row_data = view.get_row_data(self.admission.uuid)

        # Last UCL registration
        self.assertStrEqual(row_data[13], 'non')
        self.assertStrEqual(row_data[14], '')

        # High school diploma
        self.assertStrEqual(row_data[20], str(GotDiploma.NO.value))
        self.assertStrEqual(row_data[21], '')

        # Academic experience
        self.assertStrEqual(row_data[25], '2010-2011 : Computer science 2, Institute')

        # Billing address
        self.assertStrEqual(row_data[39], f'University street 2 - 1348 Louvain-la-Neuve - {self.address_country_name}')

    def test_last_experience_with_diploma(self):
        # No experience
        experiences = []
        experience = ContinuingAdmissionListExcelExportView._get_last_academic_experience(experiences)

        self.assertIsNone(experience)

        # One experience
        experiences = [
            ExperienceAcademiqueDTOFactory(
                a_obtenu_diplome=False,
                annees=[AnneeExperienceAcademiqueDTOFactory(annee=2020)],
                cycle_formation='',
            )
        ]

        experience = ContinuingAdmissionListExcelExportView._get_last_academic_experience(experiences)

        self.assertIsNone(experience)

        experiences = [
            ExperienceAcademiqueDTOFactory(
                a_obtenu_diplome=True,
                annees=[AnneeExperienceAcademiqueDTOFactory(annee=2020)],
                cycle_formation='',
            )
        ]

        experience = ContinuingAdmissionListExcelExportView._get_last_academic_experience(experiences)

        self.assertEqual(experience, experiences[0])

        # This is the graduating experience
        experiences = [
            ExperienceAcademiqueDTOFactory(
                a_obtenu_diplome=True,
                annees=[AnneeExperienceAcademiqueDTOFactory(annee=2020)],
                cycle_formation='',
            ),
            ExperienceAcademiqueDTOFactory(
                a_obtenu_diplome=False,
                annees=[AnneeExperienceAcademiqueDTOFactory(annee=2021)],
                cycle_formation='',
            ),
        ]

        experience = ContinuingAdmissionListExcelExportView._get_last_academic_experience(experiences)

        self.assertEqual(experience, experiences[0])

        # This is the latest experience
        experiences = [
            ExperienceAcademiqueDTOFactory(
                a_obtenu_diplome=True,
                annees=[AnneeExperienceAcademiqueDTOFactory(annee=2020)],
                cycle_formation='',
            ),
            ExperienceAcademiqueDTOFactory(
                a_obtenu_diplome=True,
                annees=[AnneeExperienceAcademiqueDTOFactory(annee=2021)],
                cycle_formation='',
            ),
        ]

        experience = ContinuingAdmissionListExcelExportView._get_last_academic_experience(experiences)

        self.assertEqual(experience, experiences[1])

        # This is the experience of the highest cycle
        experiences = [
            ExperienceAcademiqueDTOFactory(
                a_obtenu_diplome=True,
                annees=[AnneeExperienceAcademiqueDTOFactory(annee=2020)],
                cycle_formation=Cycle.THIRD_CYCLE.name,
            ),
            ExperienceAcademiqueDTOFactory(
                a_obtenu_diplome=True,
                annees=[AnneeExperienceAcademiqueDTOFactory(annee=2020)],
                cycle_formation=Cycle.SECOND_CYCLE.name,
            ),
        ]

        experience = ContinuingAdmissionListExcelExportView._get_last_academic_experience(experiences)

        self.assertEqual(experience, experiences[0])

        experiences = [
            ExperienceAcademiqueDTOFactory(
                a_obtenu_diplome=True,
                annees=[AnneeExperienceAcademiqueDTOFactory(annee=2020)],
                cycle_formation='',
            ),
            ExperienceAcademiqueDTOFactory(
                a_obtenu_diplome=True,
                annees=[AnneeExperienceAcademiqueDTOFactory(annee=2020)],
                cycle_formation=Cycle.FIRST_CYCLE.name,
            ),
        ]

        experience = ContinuingAdmissionListExcelExportView._get_last_academic_experience(experiences)

        self.assertEqual(experience, experiences[1])

    def test_get_non_academic_activities_columns(self):
        # No experience
        experiences = []

        columns = ContinuingAdmissionListExcelExportView._get_non_academic_activities_columns(experiences)

        self.assertEqual(columns[0], '')
        self.assertEqual(columns[1], '')
        self.assertEqual(columns[2], '')
        self.assertEqual(columns[3], '')

        experiences = [
            # Current activities
            # > Work
            ExperienceNonAcademiqueDTOFactory(
                type=ActivityType.WORK.name,
                date_debut=datetime.date(2023, 1, 1),
                date_fin=datetime.date(2023, 1, 31),
                fonction='Manager',
                secteur=ActivitySector.PRIVATE.name,
                employeur='UCL',
            ),
            # > Internship
            ExperienceNonAcademiqueDTOFactory(
                type=ActivityType.INTERNSHIP.name,
                date_debut=datetime.date(2023, 1, 1),
                date_fin=datetime.date(2023, 1, 31),
            ),
            # > Other activity
            ExperienceNonAcademiqueDTOFactory(
                type=ActivityType.OTHER.name,
                date_debut=datetime.date(2023, 1, 1),
                date_fin=datetime.date(2023, 1, 31),
                autre_activite='Other activity',
            ),
            # Past activities
            # > Other activity
            ExperienceNonAcademiqueDTOFactory(
                type=ActivityType.OTHER.name,
                date_debut=datetime.date(2022, 1, 1),
                date_fin=datetime.date(2022, 3, 31),
                autre_activite='Other activity',
            ),
            # > Internship
            ExperienceNonAcademiqueDTOFactory(
                type=ActivityType.VOLUNTEERING.name,
                date_debut=datetime.date(2022, 1, 1),
                date_fin=datetime.date(2022, 2, 28),
            ),
            # > Work
            ExperienceNonAcademiqueDTOFactory(
                type=ActivityType.WORK.name,
                date_debut=datetime.date(2022, 1, 1),
                date_fin=datetime.date(2022, 1, 31),
                fonction='Manager 0',
                secteur=ActivitySector.PRIVATE.name,
                employeur='UCL 0',
            ),
        ]
        columns = ContinuingAdmissionListExcelExportView._get_non_academic_activities_columns(experiences)

        self.assertEqual(columns[0], f'Manager\n{ActivityType.INTERNSHIP.value}\nOther activity')
        self.assertEqual(columns[1], 'UCL\n\n')
        self.assertEqual(columns[2], f'{ActivitySector.PRIVATE.value}\n\n')
        self.assertEqual(
            columns[3],
            f'01/2022-03/2022 : Other activity'
            f'\n01/2022-02/2022 : {ActivityType.VOLUNTEERING.value}'
            f'\n01/2022 : Manager 0 à UCL 0',
        )

    def test_export_configuration(self):
        candidate = PersonFactory()
        campus = CampusFactory()
        filters = str(
            {
                'annee_academique': 2022,
                'edition': [ChoixEdition.DEUX.name, ChoixEdition.TROIS.name],
                'numero': self.admission.reference,
                'matricule_candidat': self.admission.candidate.global_id,
                'etats': [ChoixStatutPropositionContinue.EN_BROUILLON.name],
                'facultes': 'ENT',
                'types_formation': [
                    TrainingType.UNIVERSITY_SECOND_CYCLE_CERTIFICATE.name,
                    TrainingType.UNIVERSITY_FIRST_CYCLE_CERTIFICATE.name,
                ],
                'sigles_formations': [self.admission.training.acronym],
                'inscription_requise': True,
                'injection_epc_en_erreur': True,
                'paye': False,
                'marque_d_interet': True,
                'mode_filtres_etats_checklist': ModeFiltrageChecklist.INCLUSION.name,
                'filtres_etats_checklist': {
                    OngletsChecklistContinue.decision.name: ['A_TRAITER', 'A_VALIDER'],
                },
                'demandeur': str(self.sic_management_user.person.uuid),
            }
        )

        view = ContinuingAdmissionListExcelExportView()
        workbook = Workbook()
        worksheet: Worksheet = workbook.create_sheet()

        view.customize_parameters_worksheet(
            worksheet=worksheet,
            person=self.sic_management_user.person,
            filters=filters,
        )

        names, values = list(worksheet.iter_cols(values_only=True))
        self.assertEqual(len(names), 17)
        self.assertEqual(len(values), 17)

        # Check the names of the parameters
        self.assertStrEqual(names[0], _('Creation date'))
        self.assertStrEqual(names[1], pgettext('masculine', 'Created by'))
        self.assertStrEqual(names[2], _('Description'))
        self.assertStrEqual(names[3], _('Year'))
        self.assertStrEqual(names[4], _('Edition'))
        self.assertStrEqual(names[5], _('Application numero'))
        self.assertStrEqual(names[6], _('Last name / First name / Email / NOMA'))
        self.assertStrEqual(names[7], _('Application status'))
        self.assertStrEqual(names[8], _('Faculty'))
        self.assertStrEqual(names[9], _('Course type'))
        self.assertStrEqual(names[10], pgettext('admission', 'Course'))
        self.assertStrEqual(names[11], _('Registration required'))
        self.assertStrEqual(names[12], _('Injection in error'))
        self.assertStrEqual(names[13], _('Paid'))
        self.assertStrEqual(names[14], _('Interested mark'))
        self.assertStrEqual(names[15], _('Include or exclude the checklist filters'))
        self.assertStrEqual(names[16], _('Checklist filters'))

        # Check the values of the parameters
        self.assertStrEqual(values[0], '3 Janvier 2023')
        self.assertStrEqual(values[1], self.sic_management_user.person.full_name)
        self.assertStrEqual(values[2], _('Export') + ' - Admissions')
        self.assertStrEqual(values[3], '2022')
        self.assertStrEqual(values[4], "['2', '3']")
        self.assertStrEqual(values[5], str(self.admission.reference))
        self.assertStrEqual(values[6], self.admission.candidate.full_name)
        self.assertStrEqual(values[7], f"['{ChoixStatutPropositionContinue.EN_BROUILLON.value}']")
        self.assertStrEqual(values[8], 'ENT')
        self.assertStrEqual(
            values[9],
            f"['{TrainingType.UNIVERSITY_SECOND_CYCLE_CERTIFICATE.value}', "
            f"'{TrainingType.UNIVERSITY_FIRST_CYCLE_CERTIFICATE.value}']",
        )
        self.assertStrEqual(values[10], f"['{self.admission.training.acronym}']")
        self.assertStrEqual(values[11], 'oui')
        self.assertStrEqual(values[12], 'oui')
        self.assertStrEqual(values[13], 'non')
        self.assertStrEqual(values[14], 'oui')
        self.assertStrEqual(values[15], str(ModeFiltrageChecklist.INCLUSION.value))
        self.assertStrEqual(
            values[16],
            str(
                {
                    OngletsChecklistContinue.decision.value: [
                        _('To be processed'),
                        _('To validate IUFC'),
                    ],
                }
            ),
        )


@freezegun.freeze_time('2023-01-03')
@override_settings(WAFFLE_CREATE_MISSING_SWITCHES=False)
class DoctorateAdmissionListExcelExportViewTestCase(QueriesAssertionsMixin, TestCase):
    def assertStrEqual(self, first, second, msg=None):
        self.assertIsInstance(first, str)
        self.assertIsInstance(second, str)
        super().assertEqual(first, second, msg)

    @classmethod
    def setUpTestData(cls):
        cls.factory = RequestFactory()

        # Users
        cls.user = User.objects.create_user(
            username='john_doe',
            password='top_secret',
        )

        cls.sic_management_user = SicManagementRoleFactory().person.user
        cls.other_sic_manager = SicManagementRoleFactory().person

        # Academic years
        cls.academic_years = AcademicYearFactory.produce(base_year=2023, number_past=2, number_future=2)

        # Entities
        faculty_entity = EntityFactory()
        EntityVersionFactory(
            entity=faculty_entity,
            acronym='ABCDEF',
            entity_type=EntityType.DOCTORAL_COMMISSION.name,
        )

        cls.first_entity = EntityFactory()
        EntityVersionFactory(
            entity=cls.first_entity,
            acronym='GHIJK',
            entity_type=EntityType.SCHOOL.name,
            parent=faculty_entity,
        )

        cls.candidate = PersonFactory(
            first_name="John",
            last_name="Doe",
            email='john.doe@example.be',
        )
        cls.student = StudentFactory(
            person=cls.candidate,
            registration_id='01234567',
        )
        cls.doctorate_committee_member = DoctorateCommitteeMemberRoleFactory().person.user

        cls.default_params = {
            'annee_academique': 2022,
            'taille_page': 10,
            'demandeur': str(cls.sic_management_user.person.uuid),
        }

        # Targeted url
        cls.url = reverse('admission:excel-exports:doctorate-admissions-list')
        cls.list_url = reverse('admission:doctorate:cdd:list')

    def test_export_user_without_person(self):
        self.client.force_login(user=self.user)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 403)

    def test_export_candidate_user(self):
        self.client.force_login(user=self.candidate.user)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 403)

    def test_export_doctorate_committee_member(self):
        self.client.force_login(user=self.doctorate_committee_member)

        response = self.client.get(self.url)

        self.assertRedirects(response, expected_url=self.list_url, fetch_redirect_response=False)

    def test_export_sic_management_user(self):
        self.client.force_login(user=self.sic_management_user)

        response = self.client.get(self.url)
        self.assertRedirects(response, expected_url=self.list_url, fetch_redirect_response=False)

        response = self.client.get(
            self.url,
            **{"HTTP_HX-Request": 'true'},
        )

        self.assertEqual(response.status_code, 200)

    def test_export_with_sic_management_user_without_filters_doesnt_plan_the_export(self):
        self.client.force_login(user=self.sic_management_user)

        response = self.client.get(self.url)

        self.assertRedirects(response, expected_url=self.list_url, fetch_redirect_response=False)

        task = AsyncTask.objects.filter(person=self.sic_management_user.person).first()
        self.assertIsNone(task)

        export = Export.objects.filter(person=self.sic_management_user.person).first()
        self.assertIsNone(export)

    def test_export_with_sic_management_user_with_filters_plans_the_export(self):
        self.client.force_login(user=self.sic_management_user)

        response = self.client.get(self.url, data=self.default_params)

        self.assertRedirects(response, expected_url=self.list_url, fetch_redirect_response=False)

        task = AsyncTask.objects.filter(person=self.sic_management_user.person).first()
        self.assertIsNotNone(task)
        self.assertEqual(task.name, _('Admission applications export'))
        self.assertEqual(task.description, _('Excel export of admission applications'))
        self.assertEqual(task.state, TaskState.PENDING.name)

        export = Export.objects.filter(job_uuid=task.uuid).first()
        self.assertIsNotNone(export)
        self.assertEqual(
            export.called_from_class,
            'admission.views.excel_exports.DoctorateAdmissionListExcelExportView',
        )
        self.assertEqual(export.file_name, 'export-des-demandes-dadmission')
        self.assertEqual(export.type, ExportTypes.EXCEL.name)

        filters = ast.literal_eval(export.filters)
        self.assertEqual(filters.get('annee_academique'), self.default_params.get('annee_academique'))
        self.assertEqual(filters.get('demandeur'), self.default_params.get('demandeur'))

    def test_export_with_sic_management_user_with_filters_and_asc_ordering(self):
        self.client.force_login(user=self.sic_management_user)

        # With asc ordering
        response = self.client.get(self.url, data={**self.default_params, 'o': 'numero_demande'})

        self.assertRedirects(response, expected_url=self.list_url, fetch_redirect_response=False)

        task = AsyncTask.objects.filter(person=self.sic_management_user.person).first()
        self.assertIsNotNone(task)

        export = Export.objects.filter(job_uuid=task.uuid).first()
        self.assertIsNotNone(export)

        filters = ast.literal_eval(export.filters)
        self.assertEqual(filters.get('annee_academique'), self.default_params.get('annee_academique'))
        self.assertEqual(filters.get('demandeur'), self.default_params.get('demandeur'))
        self.assertEqual(filters.get('tri_inverse'), False)
        self.assertEqual(filters.get('champ_tri'), 'numero_demande')

    def test_export_with_sic_management_user_with_filters_and_desc_ordering(self):
        self.client.force_login(user=self.sic_management_user)

        # With asc ordering
        response = self.client.get(self.url, data={**self.default_params, 'o': '-numero_demande'})

        self.assertRedirects(response, expected_url=self.list_url, fetch_redirect_response=False)

        task = AsyncTask.objects.filter(person=self.sic_management_user.person).first()
        self.assertIsNotNone(task)

        export = Export.objects.filter(job_uuid=task.uuid).first()
        self.assertIsNotNone(export)

        filters = ast.literal_eval(export.filters)
        self.assertEqual(filters.get('annee_academique'), self.default_params.get('annee_academique'))
        self.assertEqual(filters.get('demandeur'), self.default_params.get('demandeur'))
        self.assertEqual(filters.get('tri_inverse'), True)
        self.assertEqual(filters.get('champ_tri'), 'numero_demande')

    def test_export_content(self):
        view = DoctorateAdmissionListExcelExportView()
        header = view.get_header()

        admission = DoctorateAdmissionFactory(
            candidate=self.candidate,
            status=ChoixStatutPropositionDoctorale.CONFIRMEE.name,
            training__management_entity=self.first_entity,
            training__acronym="ZEBU0",
            training__education_group_type__name=TrainingType.PHD.name,
            submitted_at=datetime.datetime(2023, 1, 1),
            training__academic_year=self.academic_years[1],
            determined_academic_year=self.academic_years[2],
            modified_at=datetime.datetime(2023, 1, 3),
            last_update_author=self.candidate,
            cotutelle=None,
            type=ChoixTypeAdmission.ADMISSION.name,
            checklist={
                'initial': {},
                'current': {
                    'decision_sic': {'statut': 'INITIAL_CANDIDAT'},
                    'decision_cdd': {'statut': 'GEST_EN_COURS'},
                },
            },
            project_title='P3',
            with_thesis_institute=True,
        )

        # Add academic experiences
        educational_experience = EducationalExperienceFactory(
            obtained_diploma=True,
            person=admission.candidate,
            expected_graduation_date=datetime.date(2022, 6, 30),
        )

        educational_experience_year_1 = EducationalExperienceYearFactory(
            educational_experience=educational_experience,
            acquired_credit_number=15,
            academic_year__year=2020,
        )

        educational_experience_year_2 = EducationalExperienceYearFactory(
            educational_experience=educational_experience,
            acquired_credit_number=12.5,
            academic_year__year=2021,
        )

        valuation = AdmissionEducationalValuatedExperiencesFactory(
            baseadmission=admission,
            educationalexperience=educational_experience,
        )

        internal_experience = InscriptionProgrammeAnnuelFactory(
            programme_cycle__etudiant__person=admission.candidate,
            programme_cycle__decision=DecisionResultatCycle.GRANDE_DISTINCTION.name,
            programme_cycle__date_decision=datetime.date(2023, 6, 30),
            programme_cycle__credits_acquis_de_charge=29.2,
            programme__offer__title='Biology',
        )

        # Add members of the supervision group
        supervisor = ExternalPromoterFactory(is_reference_promoter=True)
        ca_member = CaMemberFactory(process=supervisor.process)
        admission.supervision_group = supervisor.process
        admission.save()

        results: List[DemandeDoctoraleRechercheDTO] = message_bus_instance.invoke(
            ListerDemandesDoctoralesQuery(
                numero=admission.reference,
                avec_experiences_academiques_reussies=True,
                avec_acteurs_groupe_supervision=True,
            )
        )

        self.assertEqual(len(results), 1)

        result = results[0]

        row_data = view.get_row_data(result)

        self.assertEqual(len(header), len(row_data))

        self.assertStrEqual(row_data[0], result.numero_demande)
        self.assertStrEqual(row_data[1], f'{result.nom_candidat}, {result.prenom_candidat} ({result.noma_candidat})')
        self.assertStrEqual(row_data[2], result.nom_pays_nationalite_candidat)
        self.assertStrEqual(row_data[3], result.code_bourse)
        self.assertStrEqual(row_data[4], f'{result.sigle_formation} - {result.intitule_formation}')
        # Academic record
        self.assertStrEqual(
            row_data[5],
            'Computer science - Institute - 2022-06-30 - 27.5 ECTS - Grande distinction (80-89%)\n'
            'Biology - UCLouvain - 2023-06-30 - 29.2 ECTS - Grande distinction',
        )
        # Supervisors
        self.assertStrEqual(
            row_data[6],
            f'{supervisor.last_name} {supervisor.first_name} ({supervisor.institute}, {supervisor.country.name})',
        )
        # Supervision committee members
        self.assertStrEqual(
            row_data[7],
            f'{ca_member.person.last_name} {ca_member.person.first_name} '
            f'(UCLouvain, {ca_member.person.country_of_citizenship.name})',
        )
        self.assertStrEqual(row_data[8], f'{admission.thesis_institute.title} ({admission.thesis_institute.acronym})')
        self.assertStrEqual(row_data[9], 'non')
        self.assertStrEqual(row_data[10], '')
        self.assertStrEqual(row_data[11], 'P3')
        self.assertStrEqual(row_data[12], str(ChoixStatutPropositionDoctorale.CONFIRMEE.value))
        self.assertStrEqual(row_data[13], _('Taken in charge'))
        self.assertStrEqual(row_data[14], _('To be processed'))
        self.assertStrEqual(row_data[15], '2023/01/01')
        self.assertStrEqual(row_data[16], '2023/01/03')
        self.assertStrEqual(
            row_data[17],
            f'{result.nom_auteur_derniere_modification}, {result.prenom_auteur_derniere_modification[:1]}',
        )

        # Check specific values
        admission.submitted_at = None
        admission.type = ChoixTypeAdmission.PRE_ADMISSION.name
        admission.cotutelle = True
        admission.save(update_fields=['submitted_at', 'type', 'cotutelle'])

        results: List[DemandeDoctoraleRechercheDTO] = message_bus_instance.invoke(
            ListerDemandesDoctoralesQuery(numero=admission.reference)
        )

        self.assertEqual(len(results), 1)

        result = results[0]

        row_data = view.get_row_data(result)

        self.assertEqual(len(header), len(row_data))

        self.assertStrEqual(row_data[9], 'oui')
        self.assertStrEqual(row_data[10], 'oui')
        self.assertStrEqual(row_data[15], '')

        admission.cotutelle = False
        admission.save(update_fields=['cotutelle'])

        results: List[DemandeDoctoraleRechercheDTO] = message_bus_instance.invoke(
            ListerDemandesDoctoralesQuery(numero=admission.reference)
        )

        self.assertEqual(len(results), 1)

        result = results[0]

        row_data = view.get_row_data(result)

        self.assertEqual(len(header), len(row_data))

        self.assertStrEqual(row_data[10], 'non')

    def test_export_configuration(self):
        country = CountryFactory()
        promoter = PromoterFactory()
        scholarship = DoctorateScholarshipFactory()
        admission = DoctorateAdmissionFactory(
            candidate=self.candidate,
            status=ChoixStatutPropositionDoctorale.CONFIRMEE.name,
            training__management_entity=self.first_entity,
            training__acronym="ZEBU0",
            training__education_group_type__name=TrainingType.PHD.name,
            submitted_at=datetime.datetime(2023, 1, 1),
            training__academic_year=self.academic_years[1],
            determined_academic_year=self.academic_years[2],
            modified_at=datetime.datetime(2023, 1, 3),
            last_update_author=self.candidate,
            cotutelle=None,
            type=ChoixTypeAdmission.ADMISSION.name,
        )

        filters = str(
            {
                'annee_academique': 2022,
                'numero': admission.reference,
                'matricule_candidat': admission.candidate.global_id,
                'nationalite': country.iso_code,
                'etats': [
                    ChoixStatutPropositionDoctorale.EN_BROUILLON.name,
                    ChoixStatutPropositionDoctorale.CONFIRMEE.name,
                ],
                'type': ChoixTypeAdmission.ADMISSION.name,
                'cdds': 'GHIJK',
                'commission_proximite': ChoixCommissionProximiteCDSS.BCGIM.name,
                'sigles_formations': ['ZEBU0'],
                'id_promoteur': json.dumps(
                    {
                        'global_id': promoter.person.global_id,
                        'last_name': promoter.person.last_name,
                        'first_name': promoter.person.first_name,
                    }
                ),
                'type_financement': ChoixTypeFinancement.SEARCH_SCHOLARSHIP.name,
                'bourse_recherche': str(scholarship.uuid),
                'cotutelle': True,
                'fnrs_fria_fresh': True,
                'date_soumission_debut': '2020-01-01',
                'date_soumission_fin': '2020-01-02',
                'mode_filtres_etats_checklist': ModeFiltrageChecklist.INCLUSION.name,
                'filtres_etats_checklist': {},
                'indicateur_tableau_bord': IndicateurTableauBordEnum.ADMISSION_DOSSIER_SOUMIS.name,
                'demandeur': str(self.sic_management_user.person.uuid),
            }
        )

        view = DoctorateAdmissionListExcelExportView()
        workbook = Workbook()
        worksheet: Worksheet = workbook.create_sheet()

        view.customize_parameters_worksheet(
            worksheet=worksheet,
            person=self.sic_management_user.person,
            filters=filters,
        )

        names, values = list(worksheet.iter_cols(values_only=True))

        self.assertEqual(len(names), 22)
        self.assertEqual(len(values), 22)

        # Check the names of the parameters
        self.assertStrEqual(names[0], _('Creation date'))
        self.assertStrEqual(names[1], pgettext('masculine', 'Created by'))
        self.assertStrEqual(names[2], _('Description'))
        self.assertStrEqual(names[3], _('Year'))
        self.assertStrEqual(names[4], _('Application numero'))
        self.assertStrEqual(names[5], _('Last name / First name / Email / NOMA'))
        self.assertStrEqual(names[6], _('Nationality'))
        self.assertStrEqual(names[7], _('Application status'))
        self.assertStrEqual(names[8], pgettext('doctorate-filter', 'Admission type'))
        self.assertStrEqual(names[9], _('Doctoral commissions'))
        self.assertStrEqual(names[10], _('Proximity commission'))
        self.assertStrEqual(names[11], pgettext('admission', 'Courses'))
        self.assertStrEqual(names[12], pgettext('gender', 'Supervisor'))
        self.assertStrEqual(names[13], _('Funding type'))
        self.assertStrEqual(names[14], _('Research scholarship'))
        self.assertStrEqual(names[15], _('Cotutelle'))
        self.assertStrEqual(names[16], _('FNRS, FRIA, FRESH, CSC'))
        self.assertStrEqual(names[17], _('Submitted from'))
        self.assertStrEqual(names[18], _('Submitted until'))
        self.assertStrEqual(names[19], _('Include or exclude the checklist filters'))
        self.assertStrEqual(names[20], _('Checklist filters'))
        self.assertStrEqual(names[21], _('Dashboard indicator'))

        # Check the values of the parameters
        self.assertStrEqual(values[0], '3 Janvier 2023')
        self.assertStrEqual(values[1], self.sic_management_user.person.full_name)
        self.assertStrEqual(values[2], _('Export') + ' - Admissions')
        self.assertStrEqual(values[3], '2022')
        self.assertStrEqual(values[4], str(admission.reference))
        self.assertStrEqual(values[5], admission.candidate.full_name)
        self.assertStrEqual(values[6], country.name)
        self.assertStrEqual(
            values[7],
            f"['{ChoixStatutPropositionDoctorale.EN_BROUILLON.value}', "
            f"'{ChoixStatutPropositionDoctorale.CONFIRMEE.value}']",
        )
        self.assertStrEqual(values[8], str(ChoixTypeAdmission.ADMISSION.value))
        self.assertStrEqual(values[9], 'GHIJK')
        self.assertStrEqual(values[10], str(ChoixCommissionProximiteCDSS.BCGIM.value))
        self.assertStrEqual(values[11], "['ZEBU0']")
        self.assertStrEqual(values[12], f'{promoter.last_name}, {promoter.first_name} ({promoter.person.global_id})')
        self.assertStrEqual(values[13], str(ChoixTypeFinancement.SEARCH_SCHOLARSHIP.value))
        self.assertStrEqual(values[14], scholarship.short_name)
        self.assertStrEqual(values[15], 'oui')
        self.assertStrEqual(values[16], 'oui')
        self.assertStrEqual(values[17], '2020-01-01')
        self.assertStrEqual(values[18], '2020-01-02')
        self.assertStrEqual(values[19], str(ModeFiltrageChecklist.INCLUSION.value))
        self.assertStrEqual(values[20], '{}')
        self.assertStrEqual(
            values[21],
            '{} - {}'.format(
                pgettext_lazy('dashboard-category', 'Admission'),
                pgettext_lazy('dashboard-indicator admission', 'Submitted dossiers'),
            ),
        )

        filters = str(
            {
                'annee_academique': 2022,
                'numero': '',
                'matricule_candidat': '',
                'nationalite': '',
                'etats': [],
                'type': '',
                'cdds': '',
                'commission_proximite': '',
                'sigles_formations': [],
                'id_promoteur': '',
                'type_financement': '',
                'bourse_recherche': '',
                'cotutelle': None,
                'fnrs_fria_fresh': None,
                'date_soumission_debut': '',
                'date_soumission_fin': '',
                'mode_filtres_etats_checklist': '',
                'filtres_etats_checklist': {},
                'indicateur_tableau_bord': '',
                'demandeur': str(self.sic_management_user.person.uuid),
            }
        )

        view = DoctorateAdmissionListExcelExportView()
        workbook = Workbook()
        worksheet: Worksheet = workbook.create_sheet()

        view.customize_parameters_worksheet(
            worksheet=worksheet,
            person=self.sic_management_user.person,
            filters=filters,
        )

        names, values = list(worksheet.iter_cols(values_only=True))

        self.assertEqual(len(names), 22)
        self.assertEqual(len(values), 22)

        # Check the values of the parameters
        self.assertStrEqual(values[0], '3 Janvier 2023')
        self.assertStrEqual(values[1], self.sic_management_user.person.full_name)
        self.assertStrEqual(values[2], _('Export') + ' - Admissions')
        self.assertStrEqual(values[3], '2022')
        self.assertStrEqual(values[4], '')
        self.assertStrEqual(values[5], '')
        self.assertStrEqual(values[6], '')
        self.assertStrEqual(values[7], '[]')
        self.assertStrEqual(values[8], '')
        self.assertStrEqual(values[9], '')
        self.assertStrEqual(values[10], '')
        self.assertStrEqual(values[11], '[]')
        self.assertStrEqual(values[12], '')
        self.assertStrEqual(values[13], '')
        self.assertStrEqual(values[14], '')
        self.assertStrEqual(values[15], '')
        self.assertStrEqual(values[16], '')
        self.assertStrEqual(values[17], '')
        self.assertStrEqual(values[18], '')
        self.assertStrEqual(values[19], '')
        self.assertStrEqual(values[20], '{}')
        self.assertStrEqual(values[21], '')
