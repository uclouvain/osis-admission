# ##############################################################################
#
#  OSIS stands for Open Student Information System. It's an application
#  designed to manage the core business of higher education institutions,
#  such as universities, faculties, institutes and professional schools.
#  The core business involves the administration of students, teachers,
#  courses, programs and so on.
#
#  Copyright (C) 2015-2026 Université catholique de Louvain (http://www.uclouvain.be)
#
#  This program is free software: you can redistribute it and/or modify
#  it under the terms of the GNU General Public License as published by
#  the Free Software Foundation, either version 3 of the License, or
#  (at your option) any later version.
#
#  This program is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU General Public License for more details.
#
#  A copy of this license - GNU General Public License - is available
#  at the root of the source code of this program.  If not,
#  see http://www.gnu.org/licenses/.
#
# ##############################################################################
import datetime
import io
import uuid
from tempfile import NamedTemporaryFile
from unittest import mock

import freezegun
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import Workbook, load_workbook
from osis_notification.models import EmailNotification

from admission.auth.scope import Scope
from admission.calendar.admission_calendar import SIGLES_WITH_QUOTA
from admission.ddd.admission.formation_generale.domain.model.enums import (
    ChoixStatutPropositionGenerale,
)
from admission.models.contingente import ContingenteTraining
from admission.tests.factories.calendar import AdmissionAcademicCalendarFactory
from admission.tests.factories.contingente import ContingenteTrainingFactory
from admission.tests.factories.general_education import (
    GeneralEducationAdmissionFactory,
    GeneralEducationTrainingFactory,
)
from admission.tests.factories.roles import CentralManagerRoleFactory
from base.models.enums.academic_calendar_type import AcademicCalendarTypes
from base.tests.factories.academic_calendar import AcademicCalendarFactory


@freezegun.freeze_time('2026-01-01')
@override_settings(WAFFLE_CREATE_MISSING_SWITCHES=False)
class ContingenteManageViewTestCase(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user = CentralManagerRoleFactory(scopes=[Scope.CONTINGENTE_NON_RESIDENT.name]).person.user

        cls.first_training = GeneralEducationTrainingFactory(
            academic_year__year=2026,
            acronym=SIGLES_WITH_QUOTA[0],
        )
        cls.second_training = GeneralEducationTrainingFactory(
            academic_year__year=2026,
            acronym=SIGLES_WITH_QUOTA[1],
        )

        AdmissionAcademicCalendarFactory.produce_all_required()

        cls.contingente_training = ContingenteTrainingFactory(training=cls.second_training, places_number=1)
        cls.admission = GeneralEducationAdmissionFactory(
            candidate__birth_date=datetime.date(2000, 1, 1),
            training=cls.second_training,
            status=ChoixStatutPropositionGenerale.CONFIRMEE.name,
            is_non_resident=True,
            ares_application_number='ARES_NUMBER',
        )

    def setUp(self) -> None:
        self.file_uuid = uuid.UUID('4bdffb42-552d-415d-9e4c-725f10dce228')

        self.confirm_remote_upload_patcher = mock.patch('osis_document_components.services.confirm_remote_upload')
        patched = self.confirm_remote_upload_patcher.start()
        patched.return_value = str(self.file_uuid)
        self.addCleanup(self.confirm_remote_upload_patcher.stop)

        self.confirm_several_remote_upload_patcher = mock.patch(
            'osis_document_components.fields.FileField._confirm_multiple_upload'
        )
        patched = self.confirm_several_remote_upload_patcher.start()
        patched.side_effect = lambda _, value, __: [str(self.file_uuid)] if value else []
        self.addCleanup(self.confirm_several_remote_upload_patcher.stop)

        self.get_remote_metadata_patcher = mock.patch('osis_document_components.services.get_remote_metadata')
        patched = self.get_remote_metadata_patcher.start()
        patched.return_value = {"name": "test.pdf", "size": 1, "mimetype": "application/pdf"}
        self.addCleanup(self.get_remote_metadata_patcher.stop)

        self.get_several_remote_metadata_patcher = mock.patch(
            'osis_document_components.services.get_several_remote_metadata'
        )
        patched = self.get_several_remote_metadata_patcher.start()
        patched.return_value = {"foo": {"name": "test.pdf", "size": 1, "mimetype": "application/pdf"}}
        self.addCleanup(self.get_several_remote_metadata_patcher.stop)

        self.get_remote_token_patcher = mock.patch('osis_document_components.services.get_remote_token')
        patched = self.get_remote_token_patcher.start()
        patched.return_value = 'foobar'
        self.addCleanup(self.get_remote_token_patcher.stop)

        self.save_raw_content_remotely_patcher = mock.patch(
            'osis_document_components.services.save_raw_content_remotely'
        )
        patched = self.save_raw_content_remotely_patcher.start()
        patched.return_value = 'a-token'
        self.addCleanup(self.save_raw_content_remotely_patcher.stop)

        patcher = mock.patch('admission.exports.utils.change_remote_metadata')
        self.change_remote_metadata_patcher = patcher.start()
        self.change_remote_metadata_patcher.return_value = 'a-token'
        self.addCleanup(patcher.stop)

        patcher = mock.patch(
            'admission.infrastructure.admission.formation_generale.domain.service.contingente.get_remote_token'
        )
        patched = patcher.start()
        patched.return_value = 'foobar'
        self.addCleanup(patcher.stop)

        # Mock weasyprint
        patcher = mock.patch(
            'admission.exports.utils.get_pdf_from_template',
            return_value=b'some content',
        )
        self.get_pdf_from_template_patcher = patcher.start()
        self.addCleanup(patcher.stop)

    def test_contingente_manage_view_get(self):
        self.client.force_login(user=self.user)

        response = self.client.get(reverse('admission:contingente:index'))

        self.assertTemplateUsed(response, 'admission/general_education/contingente/manage.html')

    def test_contingente_training_manage_view_get(self):
        self.client.force_login(user=self.user)

        response = self.client.get(reverse('admission:contingente:training') + f'?training={SIGLES_WITH_QUOTA[0]}')

        self.assertTemplateUsed(response, 'admission/general_education/contingente/manage_training.html')
        self.assertTrue(ContingenteTraining.objects.filter(training=self.first_training).exists())

    def test_contingente_training_manage_view_get_unknown_training(self):
        self.client.force_login(user=self.user)

        response = self.client.get(reverse('admission:contingente:training') + f'?training=OIHGEHGQ')

        self.assertEqual(response.status_code, 404)

    def test_contingente_training_update_view_get(self):
        self.client.force_login(user=self.user)

        response = self.client.get(
            reverse('admission:contingente:training_change', kwargs={'training': SIGLES_WITH_QUOTA[1]})
        )

        self.assertRedirects(response, reverse('admission:contingente:index'))

    def test_contingente_training_update_view_post(self):
        self.client.force_login(user=self.user)

        response = self.client.post(
            reverse('admission:contingente:training_change', kwargs={'training': SIGLES_WITH_QUOTA[1]}),
            data={'places_number': 42},
            headers={'HX-Request': "true"},
        )

        self.assertTemplateUsed(response, 'admission/general_education/contingente/contingente_training_form.html')
        self.contingente_training.refresh_from_db()
        self.assertEqual(self.contingente_training.places_number, 42)

    def test_contingente_training_export_view_get(self):
        self.client.force_login(user=self.user)

        response = self.client.get(
            reverse('admission:contingente:training_export', kwargs={'training': SIGLES_WITH_QUOTA[1]}),
        )

        self.assertEqual(response.headers['content-type'], 'application/vnd.ms-excel')
        wb = load_workbook(io.BytesIO(b"".join(response.streaming_content)))
        ws = wb.active
        lines = [[cell.value for cell in line] for line in ws]
        self.assertEqual(lines[1][10], self.admission.ares_application_number)

    def test_contingente_training_import_view_post(self):
        self.client.force_login(user=self.user)

        headers = [
            "NOM",
            "PREMIER PRENOM",
            "AUTRE PRENOM",
            "SEXE",
            "LIEU DE NAISSANCE",
            "DATE DE NAISSANCE",
            "NATIONALITE",
            "NUMERO DE LA PIECE D'IDENTITE",
            "INSTITUTION D'ENSEIGNEMENT(HAUTE ECOLE ou UNIVERSITE)",
            "GRADE CHOISI",
            "N° DOSSIER",
            "N° HUISSIER",
            "SCEAU HUISSIER",
        ]
        line = [
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "VETE",
            self.admission.ares_application_number,
            "1",
            "",
        ]
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        ws.append(line)
        excel_file = NamedTemporaryFile()
        wb.save(excel_file.name)
        excel_file.seek(0)
        excel_file.name = 'file.xlsx'

        response = self.client.post(
            reverse('admission:contingente:training_import', kwargs={'training': SIGLES_WITH_QUOTA[1]}),
            data={'import_file': excel_file},
            headers={'HX-Request': "true"},
        )
        self.assertTemplateUsed(response, "admission/general_education/contingente/manage_training_import.html")
        self.admission.refresh_from_db()
        self.contingente_training.refresh_from_db()
        self.assertEqual(self.contingente_training.last_import_at, datetime.datetime.now())
        self.assertEqual(self.contingente_training.last_import_by, self.user.person)
        self.assertEqual(self.admission.draw_number, 1)

    def test_contingente_bulk_notification_post(self):
        self.client.force_login(user=self.user)

        response = self.client.post(
            reverse('admission:contingente:training_bulk_notification', kwargs={'training': SIGLES_WITH_QUOTA[1]}),
            data={},
            headers={'HX-Request': "true"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'admission/general_education/contingente/manage_bulk_notification.html')
        self.admission.refresh_from_db()
        self.assertTrue(self.admission.non_resident_limited_enrolment_acceptance_file)
        self.assertEqual(EmailNotification.objects.count(), 1)
        email_notification: EmailNotification = EmailNotification.objects.first()
        self.assertEqual(email_notification.person, self.admission.candidate)
