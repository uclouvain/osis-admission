# ##############################################################################
#
#  OSIS stands for Open Student Information System. It's an application
#  designed to manage the core business of higher education institutions,
#  such as universities, faculties, institutes and professional schools.
#  The core business involves the administration of students, teachers,
#  courses, programs and so on.
#
#  Copyright (C) 2015-2026 Université catholique de Louvain (http://www.uclouvain.be)
#
#  This program is free software: you can redistribute it and/or modify
#  it under the terms of the GNU General Public License as published by
#  the Free Software Foundation, either version 3 of the License, or
#  (at your option) any later version.
#
#  This program is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU General Public License for more details.
#
#  A copy of this license - GNU General Public License - is available
#  at the root of the source code of this program.  If not,
#  see http://www.gnu.org/licenses/.
#
# ##############################################################################
import unicodedata

from django.contrib import messages
from django.contrib.auth.mixins import PermissionRequiredMixin
from django.core.files.temp import NamedTemporaryFile
from django.db import transaction
from django.http import FileResponse, Http404
from django.shortcuts import get_object_or_404, redirect
from django.utils import timezone
from django.utils.functional import cached_property
from django.utils.translation import gettext_lazy as _
from django.views import View
from django.views.generic import FormView, TemplateView, UpdateView
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils.exceptions import InvalidFileException

from admission.calendar.admission_calendar import SIGLES_WITH_QUOTA
from admission.ddd.admission.formation_generale.commands import (
    NotifierEnLotFormationContingenteCommand,
    RecupererAdmissionContingenteNonResidentANotifierQuery,
)
from admission.ddd.admission.formation_generale.domain.model.enums import (
    ChoixStatutPropositionGenerale,
)
from admission.forms.admission.contingente import (
    ContingenteTrainingForm,
    ContingenteTrainingImportForm,
)
from admission.models import GeneralEducationAdmission
from admission.models.contingente import ContingenteTraining
from admission.utils import (
    add_close_modal_into_htmx_response,
    add_messages_into_htmx_response,
)
from base.ddd.utils.business_validator import MultipleBusinessExceptions
from base.models.academic_calendar import AcademicCalendar
from base.models.education_group_year import EducationGroupYear
from base.models.enums.academic_calendar_type import AcademicCalendarTypes
from base.utils.utils import format_academic_year

__namespace__ = 'contingente'

__all__ = [
    "ContingenteManageView",
    "ContingenteTrainingManageView",
    "ContingenteTrainingUpdateView",
    "ContingenteTrainingExportView",
    "ContingenteTrainingImportView",
    "ContingenteBulkNotificationView",
]

from infrastructure.messages_bus import message_bus_instance

INSTITUTION_UCL_EXPORT = 'UCLouvain'


class ContingenteMixin(PermissionRequiredMixin):
    @cached_property
    def academic_year(self):
        today = timezone.now().date()
        return AcademicCalendar.objects.get(
            reference=AcademicCalendarTypes.GENERAL_EDUCATION_ENROLLMENT.name,
            start_date__lte=today,
            end_date__gte=today,
        ).data_year


class ContingenteBulkNotificationMixin:

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        admissions = message_bus_instance.invoke(
            RecupererAdmissionContingenteNonResidentANotifierQuery(
                sigle_formation=self.request.GET.get('training'),
                annee_formation=self.academic_year.year,
            )
        )

        context['training'] = self.request.GET.get('training')
        context['academic_year'] = format_academic_year(self.academic_year, short=True)
        context['bulk_notification_admissions'] = admissions
        context['bulk_notification_deadline_calendar'] = AcademicCalendar.objects.filter(
            reference=AcademicCalendarTypes.ADMISSION_NON_RESIDENT_QUOTA_ANSWER_DEADLINE.name,
            data_year=self.academic_year,
        ).first()
        context['are_admissions_already_notified'] = any(
            admission.notification_acceptation is not None for admission in admissions
        )
        return context


class ContingenteManageView(ContingenteMixin, TemplateView):
    permission_required = 'admission.view_contingente_management'
    template_name = 'admission/general_education/contingente/manage.html'
    urlpatterns = {'index': ''}

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['academic_year'] = format_academic_year(self.academic_year, short=True)
        context['drawing_calendar'] = AcademicCalendar.objects.filter(
            reference=AcademicCalendarTypes.ADMISSION_NON_RESIDENT_QUOTA_RESULT_PUBLICATION.name,
            data_year=self.academic_year,
        ).first()
        context['publication_calendar'] = AcademicCalendar.objects.filter(
            reference=AcademicCalendarTypes.ADMISSION_NON_RESIDENT_QUOTA_RESULT_PUBLICATION.name,
            data_year=self.academic_year,
        ).first()
        context['trainings'] = SIGLES_WITH_QUOTA
        return context


class ContingenteTrainingManageView(ContingenteMixin, ContingenteBulkNotificationMixin, TemplateView):
    permission_required = 'admission.view_contingente_management'
    template_name = 'admission/general_education/contingente/manage_training.html'
    urlpatterns = 'training'

    def dispatch(self, request, *args, **kwargs):
        if request.GET.get('training') not in SIGLES_WITH_QUOTA:
            raise Http404
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context['academic_year'] = format_academic_year(self.academic_year, short=True)
        context['training'] = self.request.GET.get('training')
        contingente_training, _ = ContingenteTraining.objects.get_or_create(
            training=EducationGroupYear.objects.get(
                acronym=self.request.GET.get('training'), academic_year=self.academic_year
            ),
        )
        context['contingente_training'] = contingente_training
        context['contingente_training_form'] = ContingenteTrainingForm(instance=contingente_training)
        context['contingente_training_import_form'] = ContingenteTrainingImportForm()

        return context


class ContingenteTrainingUpdateView(ContingenteMixin, UpdateView):
    permission_required = 'admission.view_contingente_management'
    template_name = 'admission/general_education/contingente/contingente_training_form.html'
    urlpatterns = {'training_change': 'training/<str:training>'}
    form_class = ContingenteTrainingForm

    def dispatch(self, request, *args, **kwargs):
        if not request.htmx:
            return redirect('admission:contingente:index')
        return super().dispatch(request, *args, **kwargs)

    def get_object(self, queryset=None):
        return get_object_or_404(
            ContingenteTraining,
            training__acronym=self.kwargs['training'],
            training__academic_year=self.academic_year,
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['contingente_training_form'] = context['form']
        return context

    def form_valid(self, form):
        form.save()
        return self.render_to_response(self.get_context_data(contingente_training_form_success=True))


class ContingenteTrainingExportView(ContingenteMixin, View):
    permission_required = 'admission.view_contingente_management'
    urlpatterns = {'training_export': 'training/<str:training>/export'}

    def get_export_data(self):
        queryset = GeneralEducationAdmission.objects.select_related('candidate__country_of_citizenship').filter(
            training__acronym=self.kwargs['training'],
            training__academic_year=self.academic_year,
            is_non_resident=True,
            status=ChoixStatutPropositionGenerale.CONFIRMEE.name,
        )

        def format_name(name):
            name = unicodedata.normalize('NFKD', name.upper())
            name = u"".join([c for c in name if not unicodedata.combining(c)])
            name = name.replace('\'', ' ')
            name = name.replace('-', ' ')
            return name

        def format_date(date):
            return date.strftime('%Y%m%d')

        for admission in queryset:
            yield [
                format_name(admission.candidate.last_name),
                format_name(admission.candidate.first_name),
                format_name(admission.candidate.middle_name),
                admission.candidate.sex,
                admission.candidate.birth_place,
                format_date(admission.candidate.birth_date),
                admission.candidate.country_of_citizenship.name,
                admission.candidate.id_card_number,
                INSTITUTION_UCL_EXPORT,
                self.kwargs['training'][:4],
                admission.ares_application_number,
                '',
                '',
            ]

    def get_export_file(self):
        wb = Workbook()
        ws = wb.active

        headers = [
            "NOM",
            "PREMIER PRENOM",
            "AUTRE PRENOM",
            "SEXE",
            "LIEU DE NAISSANCE",
            "DATE DE NAISSANCE",
            "NATIONALITE",
            "NUMERO DE LA PIECE D'IDENTITE",
            "INSTITUTION D'ENSEIGNEMENT(HAUTE ECOLE ou UNIVERSITE)",
            "GRADE CHOISI",
            "N° DOSSIER",
            "N° HUISSIER",
            "SCEAU HUISSIER",
        ]

        ws.append(headers)
        for data in self.get_export_data():
            ws.append(data)

        # Style it a bit
        ws.row_dimensions[1].font = Font(bold=True)
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length * 1.2

        tmp = NamedTemporaryFile()
        wb.save(tmp.name)
        tmp.seek(0)
        return tmp

    def get(self, request, *args, **kwargs):
        export_file = self.get_export_file()
        return FileResponse(
            export_file,
            as_attachment=True,
            filename='contingente.xlsx',
            content_type='application/vnd.ms-excel',
        )


class ContingenteTrainingImportView(ContingenteMixin, FormView):
    permission_required = 'admission.view_contingente_management'
    urlpatterns = {'training_import': 'training/<str:training>/import'}
    template_name = "admission/general_education/contingente/manage_training_import.html"
    form_class = ContingenteTrainingImportForm

    HEADERS = [
        "NOM",
        "PREMIER PRENOM",
        "AUTRE PRENOM",
        "SEXE",
        "LIEU DE NAISSANCE",
        "DATE DE NAISSANCE",
        "NATIONALITE",
        "NUMERO DE LA PIECE D'IDENTITE",
        "INSTITUTION D'ENSEIGNEMENT(HAUTE ECOLE ou UNIVERSITE)",
        "GRADE CHOISI",
        "N° DOSSIER",
        "N° HUISSIER",
        "SCEAU HUISSIER",
    ]

    def dispatch(self, request, *args, **kwargs):
        if not request.htmx:
            return redirect('admission:contingente:index')
        response = super().dispatch(request, *args, **kwargs)
        add_close_modal_into_htmx_response(response)
        return response

    def get_object(self, queryset=None):
        return get_object_or_404(
            ContingenteTraining,
            training__acronym=self.kwargs['training'],
            training__academic_year=self.academic_year,
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['contingente_training'] = self.get_object()
        context['contingente_training_import_form'] = context['form']
        return context

    def import_wb(self, wb: Workbook):
        ws = wb.active
        lines = [[cell.value for cell in line] for line in ws]
        headers = lines.pop(0)
        if headers != self.HEADERS:
            raise ValueError("Headers are incorrects")
        GeneralEducationAdmission.objects.filter(training=self.contingente_training.training).update(draw_number=None)
        training = self.kwargs['training'][:4]
        for line in lines:
            if line[9] != training:
                # Line is for another training, ignore it
                continue
            if line[10] and line[11]:
                GeneralEducationAdmission.objects.filter(ares_application_number=line[10]).update(
                    draw_number=int(line[11])
                )

    def form_valid(self, form):
        self.contingente_training = self.get_object()
        try:
            wb = load_workbook(form.cleaned_data['import_file'])
            with transaction.atomic():
                self.import_wb(wb)
                self.contingente_training.last_import_at = timezone.now()
                self.contingente_training.last_import_by = self.request.user.person
                self.contingente_training.save(update_fields=['last_import_at', 'last_import_by'])
        except InvalidFileException as e:
            messages.error(self.request, _("Invalid file format ."))
        except Exception as e:
            messages.error(
                self.request, _("There was an error while importing the file: {error}.").format(error=str(e))
            )

        response = self.render_to_response(self.get_context_data())
        add_messages_into_htmx_response(self.request, response)
        return response


class ContingenteBulkNotificationView(ContingenteMixin, ContingenteBulkNotificationMixin, TemplateView):
    permission_required = 'admission.view_contingente_management'
    urlpatterns = {'training_bulk_notification': 'training/<str:training>/bulk_notification'}
    template_name = "admission/general_education/contingente/manage_bulk_notification.html"

    http_method_names = ['post']

    def post(self, request, *args, **kwargs):

        try:
            message_bus_instance.invoke(
                NotifierEnLotFormationContingenteCommand(
                    sigle_formation=self.kwargs['training'],
                    annee_formation=self.academic_year.year,
                    gestionnaire=request.user.person.global_id,
                )
            )
            extra_context = {
                'bulk_notification_success': True,
            }
        except MultipleBusinessExceptions as multiple_exceptions:
            message = multiple_exceptions.exceptions.pop().message
            extra_context = {
                'bulk_notification_error': message,
            }

        context = self.get_context_data(**kwargs)
        context.update(extra_context)
        response = self.render_to_response(context)
        add_close_modal_into_htmx_response(response)
        return response
